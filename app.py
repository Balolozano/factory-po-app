"""
Factory PO → Odoo Excel Converter (Ultra-Low Cost)
===================================================
Streamlit app that extracts purchase order data from Excel, PDF, or image files
and produces an Odoo-ready import Excel file.

Uses Claude Haiku for minimal token cost with hybrid local preprocessing.
"""

import streamlit as st
import pandas as pd
import json
import io
import os
import re
import base64
import tempfile
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="LODI — Conversor de Pedidos",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
APP_PASSWORD = os.environ.get("APP_PASSWORD", "Lodi3501")
DEFAULT_MODEL = "claude-haiku-4-5-20251001"
MAX_IMAGE_SIDE = 1568
JPEG_QUALITY = 80

# Odoo columns the AI must extract (A–E + G)
EXTRACT_COLUMNS = [
    "client_order_ref",
    "date_order",
    "commitment_date",
    "*//Cliente",
    "Producto",
    "order_line / product_uom_qty",
]

# Full Odoo template column order (for final export)
ALL_ODOO_COLUMNS = [
    "client_order_ref",
    "date_order",
    "commitment_date",
    "*//Cliente",
    "Producto",
    "Codigo Valido?",
    "order_line / product_uom_qty",
    "id",
    "partner_id / id",
    "order_line / product_id / id",
]

# Columns shown in the data editor (internal id columns are hidden from view)
DISPLAY_COLUMNS = [
    "client_order_ref",
    "date_order",
    "commitment_date",
    "*//Cliente",
    "Producto",
    "Codigo Valido?",
    "order_line / product_uom_qty",
]

# ---------------------------------------------------------------------------
# Language strings
# ---------------------------------------------------------------------------
LANG = {
    "es": {
        "page_title":       "LODI — Conversor de Pedidos",
        "toggle_btn":       "🇺🇸 English",
        "login_title":      "Conversor de Pedidos → Odoo",
        "login_sub":        "Ingresa la contraseña del equipo para continuar",
        "login_btn":        "Entrar",
        "login_error":      "Contraseña incorrecta. Inténtalo de nuevo.",
        "login_footer":     "LODI Manufacturing · Monterrey, Mexico",
        "header_title":     "Conversor de Pedidos de Compra → Odoo",
        "header_sub":       "Extrae automáticamente los datos de tus POs y genera el archivo listo para importar en Odoo",
        "upload_label":     "📂 Sube tus órdenes de compra",
        "upload_hint":      "Arrastra y suelta archivos aquí, o haz clic para seleccionar",
        "files_selected":   "{n} archivo(s) seleccionado(s)",
        "process_btn":      "⚡ Procesar Pedidos con IA",
        "prog_pre":         "Preprocesando archivos localmente…",
        "prog_file":        "Leyendo: {f}",
        "prog_claude":      "Consultando Claude AI…",
        "prog_done":        "¡Extracción completa!",
        "prog_step1":       "📄 Leyendo archivos...",
        "prog_step2":       "🤖 Analizando con IA...",
        "prog_step3":       "✅ ¡Extracción completa!",
        "warn_pdf":         "No se pudo procesar el PDF escaneado: {f}",
        "results_title":    "📋 Líneas de pedido extraídas",
        "results_sub":      "Revisa y corrige cualquier valor antes de descargar. Puedes editar celdas, agregar o eliminar filas.",
        "download_btn":     "📥 Descargar Excel para Odoo",
        "raw_expander":     "🔍 Ver análisis de la IA",
        "sb_tool":          "ℹ️ Herramienta",
        "sb_tool_desc":     "Extrae datos de órdenes de compra en Excel, PDF o imágenes y genera el archivo listo para importar en Odoo.",
        "logout_btn":       "Cerrar sesión",
        "col_qty":          "Cantidad",
        "col_date":         "Fecha Pedido",
        "col_commit":       "Fecha Compromiso",
        "col_ref":          "Ref. PO",
        "col_client":       "Cliente",
        "col_product":      "Producto",
        "col_valid":        "¿Válido?",
        "api_label":        "🔑 API Key",
        "api_active":       "✅ API Key activa",
        "lines_suffix":     "línea(s)",
        "override_title":   "📋 Completar información manualmente (opcional)",
        "override_sub":     "Si el PO no tiene estos datos, ingrésalos aquí y se aplicarán a todas las líneas.",
        "override_cliente": "Cliente",
        "override_ref":     "Ref. PO",
        "override_date":    "Fecha Pedido (YYYY-MM-DD)",
        "override_commit":  "Fecha Compromiso (YYYY-MM-DD)",
        "new_po_btn":       "🔄 Nueva PO",
        "code_choice_label": "🔀 Este PO tiene dos columnas de código — ¿cuál usar como Producto?",
        "code_lodi":        "Columna 1",
        "code_cliente":     "Columna 2",
        "step1":            "1 · Subir archivo",
        "step2":            "2 · Procesar con IA",
        "step3":            "3 · Descargar Excel",
        "formats_label":    "Formatos aceptados",
    },
    "en": {
        "page_title":       "LODI — PO Converter",
        "toggle_btn":       "🇲🇽 Español",
        "login_title":      "Purchase Order → Odoo Converter",
        "login_sub":        "Enter your team password to continue",
        "login_btn":        "Sign In",
        "login_error":      "Incorrect password. Please try again.",
        "login_footer":     "LODI Manufacturing · Monterrey, Mexico",
        "header_title":     "Purchase Order → Odoo Converter",
        "header_sub":       "Automatically extract PO data and generate a ready-to-import Odoo file",
        "upload_label":     "📂 Upload your purchase orders",
        "upload_hint":      "Drag and drop files here, or click to browse",
        "files_selected":   "{n} file(s) selected",
        "process_btn":      "⚡ Process Orders with AI",
        "prog_pre":         "Preprocessing files locally…",
        "prog_file":        "Reading: {f}",
        "prog_claude":      "Calling Claude AI…",
        "prog_done":        "Extraction complete!",
        "prog_step1":       "📄 Reading files...",
        "prog_step2":       "🤖 Analyzing with AI...",
        "prog_step3":       "✅ Extraction complete!",
        "warn_pdf":         "Could not process scanned PDF: {f}",
        "results_title":    "📋 Extracted PO Lines",
        "results_sub":      "Review and correct any values before downloading. You can edit cells, add or remove rows.",
        "download_btn":     "📥 Download Odoo-Ready Excel",
        "raw_expander":     "🔍 View AI Analysis",
        "sb_tool":          "ℹ️ About",
        "sb_tool_desc":     "Extracts data from purchase orders in Excel, PDF or images and generates a file ready to import into Odoo.",
        "logout_btn":       "Sign out",
        "col_qty":          "Quantity",
        "col_date":         "Order Date",
        "col_commit":       "Commitment Date",
        "col_ref":          "PO Ref.",
        "col_client":       "Customer",
        "col_product":      "Product",
        "col_valid":        "Valid?",
        "api_label":        "🔑 API Key",
        "api_active":       "✅ API Key active",
        "lines_suffix":     "line(s)",
        "override_title":   "📋 Fill in information manually (optional)",
        "override_sub":     "If the PO is missing any of these fields, enter them here and they will be applied to all lines.",
        "override_cliente": "Client",
        "override_ref":     "PO Reference",
        "override_date":    "Order Date (YYYY-MM-DD)",
        "override_commit":  "Commitment Date (YYYY-MM-DD)",
        "new_po_btn":       "🔄 New PO",
        "code_choice_label": "🔀 This PO has two code columns — which one to use as Product?",
        "code_lodi":        "Column 1",
        "code_cliente":     "Column 2",
        "step1":            "1 · Upload file",
        "step2":            "2 · Process with AI",
        "step3":            "3 · Download Excel",
        "formats_label":    "Accepted formats",
    },
}

if "lang" not in st.session_state:
    st.session_state.lang = "es"
T = LANG[st.session_state.lang]

# ---------------------------------------------------------------------------
# System prompt (cached via Anthropic prompt caching)
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """You are a purchase-order extraction engine for LODI Manufacturing, a factory in Monterrey, Mexico.
LODI Manufacturing is the VENDOR/SELLER — they RECEIVE purchase orders FROM their clients.
Your ONLY job: read the raw text or image of a purchase order and return a JSON array of line items.

Each object in the array MUST have exactly these keys (no extras):
{
  "client_order_ref": "PO or order reference number (string)",
  "date_order": "order date in YYYY-MM-DD format",
  "commitment_date": "promised/required delivery date in YYYY-MM-DD, or 2 days after date_order if not found",
  "cliente": "the company that ISSUED this PO — the BUYER (string)",
  "producto": "LODI's PART NUMBER or PRODUCT CODE — see CRITICAL section below",
  "producto_cliente": "the CLIENT's own code/part number if the PO has a SEPARATE column for it — else use \"\"",
  "qty": quantity as a number (integer or decimal, no text)
}

RULES:
- Documents may be in Spanish or English. Handle both perfectly.
- One JSON object per line item (a PO with 5 products → 5 objects).
- client_order_ref, date_order, commitment_date, and cliente are usually the same for every line in one PO. Repeat them on every object.
- Dates: convert ANY format (dd/mm/yyyy, mm-dd-yyyy, "15 de marzo 2026", etc.) to YYYY-MM-DD.
- qty must be a raw number. Strip units like "pzas", "kg", "pcs", etc.
- If a field is truly not found, use "" for strings and 0 for qty.
- Return ONLY the JSON array, no markdown, no explanation, no extra keys.
- Use COMPACT JSON — no spaces, no indentation, no newlines between fields. Every object on one line.

CRITICAL — cliente field:
A purchase order is sent BY the buyer TO the vendor (LODI). The document layout is typically:
  - TOP of document (letterhead/header): buyer's company name and logo → THIS is the cliente
  - "Vendor:" section: LODI's own address → NEVER use this as cliente
  - "Ship To:" section: delivery/warehouse address → NEVER use this as cliente

Rules:
- The CLIENTE is the company whose name/logo appears in the HEADER or LETTERHEAD at the very top of the document.
- Any section explicitly labeled "Vendor:", "Proveedor:", "Supplier:", or "Sell To:" contains LODI's info — IGNORE it for the cliente field.
- Any section labeled "Ship To:", "Deliver To:", "Enviar a:" is a shipping address — IGNORE it for the cliente field.
- LODI, INDUSTRIAS AUTOMOTRICES LODI, LODI S.A DE C.V, LODI Manufacturing, or any variant = VENDOR (us) — NEVER use as cliente.
- Look at the very top-left of the document: the company name there (often above their address, phone, fax) is the cliente.
- Example: if the top of the PO shows "AP Exhaust Technologies, Inc. / 300 Dixie Trail..." → cliente = "AP Exhaust Technologies, Inc."
- IMPORTANT: Some PO letterheads are images/logos and will NOT appear as text. In that case, look for other clues to identify the buyer:
  * Website URLs in the document (e.g. "apemissions.com" → "AP Exhaust Technologies", "company.com/purchase-order" → company name)
  * Email addresses (e.g. "buyer@acmecorp.com" → "ACME Corp")
  * Any "terms and conditions" links often contain the buyer's domain
  * Phone/fax labels like "AP Corp Purchasing:", "XYZ Purchasing:" reveal the buyer name
- When using a URL to infer the buyer, return the full proper company name, not just the domain.

CRITICAL — producto field (PART NUMBER only, never the description):
The "producto" field MUST contain ONLY the PART NUMBER or PRODUCT CODE — never the written description.
- Look for column/field labels like: "Part No", "Part #", "Item No", "Item Number", "Part Number",
  "Product Code", "Code", "SKU", "Ref", "Número de parte", "Clave", "Código", "No. Parte".
- The part number is a SHORT alphanumeric token: "40630", "01622-5", "P12345-A", "WDM 560075".
- The description is the LONG text explaining the part: "FLANGE: 3 BLT TRIAN", "EXHAUST PIPE 2IN", etc.
- NEVER combine them. NEVER output "40630 FLANGE: 3 BLT TRIAN" — output ONLY "40630".
- Example PO line:  1 | 40630 | FLANGE: 3 BLT TRIAN 4-3/4 X 3-1/2 X 3/8 THK | 1000 EA
  → producto = "40630"
- If ONLY a description is visible and no code exists, put the description in "producto".

CRITICAL — dual code columns (producto vs producto_cliente):
Some POs list TWO separate code columns: LODI's own code AND the client's code.
- If the PO has ONE code column → put it in "producto", leave "producto_cliente" as "".
- If the PO has TWO code columns with labels such as:
    "Vendor Part #" / "Customer Part #", "Our Code" / "Your Code",
    "Código Proveedor" / "Código Cliente", "No. Proveedor" / "No. Cliente", etc.:
  → put LODI's / vendor / supplier code in "producto"
  → put the client's / buyer / customer code in "producto_cliente"
- Both must still be PART NUMBERS only (never descriptions).
- If unsure which is which, put the shorter/numeric one in "producto" and the other in "producto_cliente".

SCANNED / IMAGE PDFs:
Some inputs will be rendered page images (photos or scans of paper POs). When processing these:
- Read all visible text including handwritten notes, stamps, and printed text.
- Pay special attention to the company logo or letterhead at the top — that is the cliente.
- Fax headers, stamps, and watermarks may contain dates — use the PURCHASE ORDER date, not a fax date.
- If a field is partially visible or blurry, extract your best reading and flag uncertainty with a "?" suffix."""

# ---------------------------------------------------------------------------
# Password gate
# ---------------------------------------------------------------------------
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    # ── Encode logo ──────────────────────────────────────────────────────────
    _login_logo_b64 = ""
    _lp = Path(__file__).parent / "lodi_logo.png"
    if _lp.exists():
        with open(_lp, "rb") as _lf:
            _login_logo_b64 = base64.b64encode(_lf.read()).decode()
    _logo_tag = (
        f'<img src="data:image/png;base64,{_login_logo_b64}" '
        'style="height:100px;object-fit:contain;pointer-events:none;user-select:none" alt="LODI">'
        if _login_logo_b64
        else '<span style="font-size:2.5rem;font-weight:900;color:#E8622A">LODI</span>'
    )

    # ── UI CHANGE: Login page — full redesign with animated mesh gradient, ──
    # ── larger card, orange top-border, improved typography, thicker separator ──
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    #MainMenu, footer, header, [data-testid="stToolbar"] { visibility: hidden; display: none; }

    html, body { font-family: 'Inter', sans-serif !important; }

    /* ── Animated mesh gradient background ── */
    @keyframes mesh-drift {
        0%   { background-position: 0% 0%, 100% 100%, 50% 50%; }
        33%  { background-position: 30% 15%, 65% 85%, 40% 60%; }
        66%  { background-position: 15% 40%, 85% 55%, 60% 30%; }
        100% { background-position: 0% 0%, 100% 100%, 50% 50%; }
    }
    .stApp {
        background-color: #080C12 !important;
        background-image:
            radial-gradient(ellipse 70% 55% at 25% -10%, rgba(232,98,42,0.22) 0%, transparent 55%),
            radial-gradient(ellipse 55% 45% at 85% 110%, rgba(245,158,11,0.1) 0%, transparent 50%),
            radial-gradient(ellipse 40% 30% at 70% 40%, rgba(232,98,42,0.06) 0%, transparent 50%) !important;
        background-size: 100% 100%;
        animation: mesh-drift 14s ease-in-out infinite;
        min-height: 100vh;
    }

    section[data-testid="stMain"] { background: transparent !important; }

    /* ── Card: larger, orange top-border, more padding ── */
    .block-container {
        max-width: 460px !important;
        margin: max(5vh, 40px) auto 2rem !important;
        padding: 3rem 3rem 2.5rem !important;
        background: #ffffff !important;
        border-radius: 20px !important;
        border-top: 3px solid #E8622A !important;
        box-shadow:
            0 0 0 1px rgba(232,98,42,0.12),
            0 24px 64px rgba(0,0,0,0.55),
            0 60px 120px rgba(0,0,0,0.3) !important;
    }

    /* ── Language button — right-aligned pill ── */
    .st-key-lang_login {
        display: flex !important;
        justify-content: flex-end !important;
        margin-bottom: 0.8rem !important;
    }
    .st-key-lang_login button {
        background: #f1f5f9 !important;
        border: 1.5px solid #e2e8f0 !important;
        color: #475569 !important;
        border-radius: 99px !important;
        font-size: 0.78rem !important;
        font-weight: 600 !important;
        padding: 4px 14px !important;
        box-shadow: none !important;
        white-space: nowrap !important;
        width: auto !important;
        min-width: 0 !important;
        letter-spacing: 0.2px;
        transition: all 0.18s cubic-bezier(0.4,0,0.2,1) !important;
    }
    .st-key-lang_login button:hover {
        background: #E8622A !important;
        border-color: #E8622A !important;
        color: #fff !important;
        box-shadow: none !important;
    }

    /* ── Password input (light card context) ── */
    [data-testid="stTextInput"] > div > div > input {
        background: #f8fafc !important;
        border: 1.5px solid #e2e8f0 !important;
        border-radius: 12px !important;
        font-size: 1rem !important;
        font-family: 'Inter', sans-serif !important;
        color: #0f172a !important;
        padding: 0.7rem 1rem 0.7rem 2.8rem !important;
        transition: border-color 0.18s, box-shadow 0.18s !important;
    }
    [data-testid="stTextInput"] > div > div > input:focus {
        border-color: #E8622A !important;
        box-shadow: 0 0 0 3px rgba(232,98,42,0.13) !important;
        background: #fff !important;
        outline: none !important;
    }

    /* Lock icon pseudo-element on the input wrapper */
    [data-testid="stTextInput"] > div > div {
        position: relative !important;
    }
    [data-testid="stTextInput"] > div > div::before {
        content: "🔒";
        position: absolute;
        left: 0.85rem;
        top: 50%;
        transform: translateY(-50%);
        font-size: 0.9rem;
        pointer-events: none;
        z-index: 1;
    }

    /* ── Sign-in button ── */
    .stFormSubmitButton > button {
        background: linear-gradient(135deg, #E8622A 0%, #f59e0b 100%) !important;
        border: none !important;
        color: #fff !important;
        font-weight: 700 !important;
        font-size: 1rem !important;
        font-family: 'Inter', sans-serif !important;
        border-radius: 12px !important;
        padding: 0.75rem !important;
        letter-spacing: 0.3px !important;
        box-shadow: 0 4px 20px rgba(232,98,42,0.38) !important;
        transition: opacity 0.18s, transform 0.15s, box-shadow 0.18s !important;
        width: 100% !important;
    }
    .stFormSubmitButton > button:hover {
        opacity: 0.9 !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 8px 28px rgba(232,98,42,0.52) !important;
    }

    /* Remove form border/bg */
    [data-testid="stForm"] {
        border: none !important;
        padding: 0 !important;
        background: transparent !important;
    }

    /* Error alert */
    [data-testid="stAlert"] {
        border-radius: 10px !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.875rem !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Language toggle (Streamlit button, right-aligned via CSS) ────────────
    if st.button(T["toggle_btn"], key="lang_login"):
        st.session_state.lang = "en" if st.session_state.lang == "es" else "es"
        st.rerun()

    # ── Logo ─────────────────────────────────────────────────────────────────
    st.markdown(
        f"<div style='text-align:center;margin-bottom:1.6rem'>{_logo_tag}</div>",
        unsafe_allow_html=True,
    )

    # ── UI CHANGE: Title uses Display scale (1.4rem/800), subtitle improved ──
    # ── Separator is now 3px with full-width fade ─────────────────────────────
    st.markdown(
        f"<div style='text-align:center;font-size:1.4rem;font-weight:800;"
        f"color:#0f172a;font-family:Inter,sans-serif;margin-bottom:0.3rem;letter-spacing:-0.3px'>"
        f"{T['login_title']}</div>"
        f"<div style='text-align:center;font-size:0.875rem;color:#64748b;"
        f"font-family:Inter,sans-serif;margin-bottom:1.5rem;font-weight:400'>{T['login_sub']}</div>"
        f"<div style='height:3px;background:linear-gradient(90deg,transparent 0%,#E8622A 25%,"
        f"#f59e0b 75%,transparent 100%);border-radius:2px;margin-bottom:1.6rem'></div>",
        unsafe_allow_html=True,
    )

    # ── Login form ────────────────────────────────────────────────────────────
    with st.form("login_form", clear_on_submit=False):
        pwd = st.text_input(
            "pwd", type="password",
            placeholder="••••••••",
            label_visibility="collapsed",
        )
        submitted = st.form_submit_button(T["login_btn"], use_container_width=True)
        if submitted:
            if pwd == APP_PASSWORD:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error(T["login_error"])

    # ── UI CHANGE: Footer — intentionally subtle at 0.75rem, #3D5166 ─────────
    st.markdown(
        f"<div style='text-align:center;font-size:0.75rem;color:#3D5166;"
        f"margin-top:1.6rem;font-family:Inter,sans-serif;letter-spacing:0.4px'>"
        f"{T['login_footer']}</div>",
        unsafe_allow_html=True,
    )
    st.stop()

# Refresh T after auth (language may have changed)
T = LANG[st.session_state.lang]

# ---------------------------------------------------------------------------
# Lazy imports (only after auth, so cold-start feels faster)
# ---------------------------------------------------------------------------
import anthropic  # noqa: E402

@st.cache_resource
def load_odoo_clients() -> list[str]:
    """Load client names from the Clientes sheet. Uses pandas for fast bulk read."""
    template_path = Path(__file__).parent / "Carga Pedidos Odoo.xlsx"
    if not template_path.exists():
        return []
    try:
        df = pd.read_excel(str(template_path), sheet_name="Clientes",
                           header=0, usecols=[0], dtype=str)
        df.columns = ["name"]
        df = df.dropna(subset=["name"])
        df["name"] = df["name"].str.strip()
        return df.loc[df["name"].str.len() > 2, "name"].tolist()
    except Exception:
        return []

# Warm up cache so it's ready when the user hits Process
load_odoo_clients()


def _normalize_client(s: str) -> str:
    """Strip legal suffixes and punctuation for fuzzy comparison."""
    s = s.lower().strip()
    s = re.sub(
        r"\b(inc\.?|llc\.?|ltd\.?|corp\.?|sa\s+de\s+cv|s\s+de\s+rl\s+de\s+cv|"
        r"sapi\s+de\s+cv|s\.?a\.?|de\s+c\.?v\.?|gmbh|plc|pty\.?|co\.?)\b",
        "", s,
    )
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def resolve_client_name(raw_name: str, api_key: str) -> str:
    """
    Map the AI-extracted client name to the closest entry in the Odoo Clientes list.

    Steps:
    1. Exact match (case-insensitive) — instant, free.
    2. Fuzzy match (rapidfuzz token_set_ratio on normalized names) to get top 5 candidates.
    3. Micro Claude call to semantically pick the best candidate.
    Returns the matched Odoo name, or the original raw_name if no good match is found.
    """
    if not raw_name or not raw_name.strip():
        return raw_name

    clients = load_odoo_clients()
    if not clients:
        return raw_name

    # ── Step 1: exact match (case-insensitive) ───────────────────────────────
    raw_lower = raw_name.strip().lower()
    for c in clients:
        if c.lower() == raw_lower:
            return c

    # ── Step 2: fuzzy candidates ─────────────────────────────────────────────
    try:
        from rapidfuzz import process as rf_process, fuzz as rf_fuzz
    except ImportError:
        return raw_name  # rapidfuzz not installed — skip matching

    norm_raw = _normalize_client(raw_name)
    norm_clients = [_normalize_client(c) for c in clients]

    candidates_raw = rf_process.extract(
        norm_raw, norm_clients, scorer=rf_fuzz.token_set_ratio,
        limit=5, score_cutoff=45,
    )
    if not candidates_raw:
        return raw_name  # nothing close enough

    # Map back to original names
    candidates = [clients[idx] for _, _, idx in candidates_raw]

    # ── Step 3: Claude picks the best semantic match ─────────────────────────
    try:
        client_api = anthropic.Anthropic(api_key=api_key)
        numbered = "\n".join(f"{i+1}. {c}" for i, c in enumerate(candidates))
        response = client_api.messages.create(
            model=DEFAULT_MODEL,
            max_tokens=80,
            messages=[{
                "role": "user",
                "content": (
                    f'A purchase order was issued by: "{raw_name}"\n\n'
                    f"Which of these Odoo client records is the same company?\n"
                    f"Consider name variations, legal suffix differences (Inc vs LLC), "
                    f"and product line name differences (e.g. Exhaust vs Emissions).\n\n"
                    f"{numbered}\n\n"
                    f"Reply with ONLY the exact name from the list above, "
                    f'or "none" if none of them are the same company.'
                ),
            }],
        )
        answer = response.content[0].text.strip().strip('"').strip("'")
        # Validate the answer is actually one of our candidates
        for c in candidates:
            if c.lower() == answer.lower():
                return c
        # Fuzzy fallback: if Claude returned a close variant of a candidate
        for c in candidates:
            if answer.lower() in c.lower() or c.lower() in answer.lower():
                return c
    except Exception:
        pass

    # Fall back to top fuzzy candidate if score is high enough
    best_score = candidates_raw[0][1]
    if best_score >= 75:
        return candidates[0]

    return raw_name  # no confident match


def get_pdfplumber():
    import pdfplumber
    return pdfplumber


def get_pillow():
    from PIL import Image
    return Image


# ---------------------------------------------------------------------------
# Helper: Extract text from files locally (before calling Claude)
# ---------------------------------------------------------------------------
def extract_text_from_excel(file_bytes: bytes, filename: str) -> str:
    """Read Excel into a compact text representation."""
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
        df = df.fillna("")
        # Compact CSV-style representation to minimise tokens
        return f"[Excel file: {filename}]\n" + df.to_csv(index=False)
    except Exception as e:
        return f"[Error reading Excel {filename}: {e}]"


def extract_text_from_pdf(file_bytes: bytes, filename: str) -> tuple[str, bool]:
    """
    Try pdfplumber text+tables first.
    Returns (text, needs_vision).  needs_vision=True means scanned/image PDF.

    Uses layout=True so side-by-side columns (e.g. Vendor vs Ship To) are
    spatially separated in the output rather than merged on the same line.

    For scanned PDFs (no text layer), we skip EasyOCR and go straight to
    Claude Vision — far more accurate for PO documents with logos and tables.
    """
    pdfplumber = get_pdfplumber()
    text_parts = []
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                # layout=True preserves column positions — critical for PO forms
                # where Vendor and Ship To sit side-by-side.
                page_text = page.extract_text(layout=True) or ""
                tables = page.extract_tables() or []
                table_text = ""
                for t in tables:
                    for row in t:
                        table_text += " | ".join(str(c) if c else "" for c in row) + "\n"
                combined = (page_text + "\n" + table_text).strip()
                if combined:
                    text_parts.append(f"--- Page {i} ---\n{combined}")
    except Exception as e:
        return f"[Error reading PDF {filename}: {e}]", False

    full_text = "\n".join(text_parts).strip()
    if len(full_text) < 40:
        # Scanned/image-only PDF — use Claude Vision for best accuracy
        # (Vision reads logos, stamps, handwriting, and complex layouts)
        return "", True
    return f"[PDF file: {filename}]\n{full_text}", False


def extract_text_from_image(file_bytes: bytes, filename: str) -> tuple[str, bool]:
    """Always send images to Claude Vision — more accurate than local OCR."""
    return "", True


def resize_image_for_vision(file_bytes: bytes) -> str:
    """Resize and compress image, return base64 JPEG."""
    Image = get_pillow()
    img = Image.open(io.BytesIO(file_bytes)).convert("RGB")
    w, h = img.size
    if max(w, h) > MAX_IMAGE_SIDE:
        scale = MAX_IMAGE_SIDE / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=JPEG_QUALITY)
    return base64.standard_b64encode(buf.getvalue()).decode()


# ---------------------------------------------------------------------------
# Helper: Render PDF pages to vision images (no system deps via PyMuPDF)
# ---------------------------------------------------------------------------
def render_pdf_for_vision(file_bytes: bytes, fname: str) -> list[tuple[str, str]]:
    """
    Convert every page of a PDF to a base64 JPEG for Claude Vision.

    Tries PyMuPDF (fitz) first — pure Python wheel, no system poppler required.
    Falls back to pdf2image+poppler if PyMuPDF is unavailable.
    Raises RuntimeError if both methods fail.
    """
    vision_images: list[tuple[str, str]] = []

    # ── Method 1: PyMuPDF (preferred — no system dependencies) ──────────────
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc[page_num]
            # 150 DPI matrix (PDF unit = 1/72 inch → scale = 150/72)
            mat = fitz.Matrix(150 / 72, 150 / 72)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            img_bytes = pix.tobytes("jpeg")
            vision_images.append((resize_image_for_vision(img_bytes), fname))
        doc.close()
        return vision_images
    except ImportError:
        pass  # PyMuPDF not installed — try next method
    except Exception:
        vision_images.clear()  # discard any partial results

    # ── Method 2: pdf2image + poppler (fallback) ─────────────────────────────
    from pdf2image import convert_from_bytes
    imgs = convert_from_bytes(file_bytes, dpi=150)
    for pg_img in imgs:
        buf = io.BytesIO()
        pg_img.save(buf, format="JPEG", quality=JPEG_QUALITY)
        vision_images.append((resize_image_for_vision(buf.getvalue()), fname))
    return vision_images


# ---------------------------------------------------------------------------
# Helper: Call Claude API
# ---------------------------------------------------------------------------
def call_claude(text_content: str, vision_images: list[tuple[str, str]]) -> str:
    """
    Send extracted text (and optional vision images) to Claude.
    Uses prompt caching for the system prompt.
    Returns raw response text.
    """
    api_key = (
        st.session_state.get("manual_api_key", "").strip()
        or os.environ.get("ANTHROPIC_API_KEY", "")
        or st.secrets.get("ANTHROPIC_API_KEY", "")
    )
    if not api_key:
        st.error("ANTHROPIC_API_KEY no configurada. Ingrésala en la barra lateral.")
        st.stop()

    client = anthropic.Anthropic(api_key=api_key)

    # Build user message content blocks
    user_blocks = []
    if text_content.strip():
        user_blocks.append({
            "type": "text",
            "text": f"Extract all PO line items from the following documents:\n\n{text_content}",
        })

    for b64_data, fname in vision_images:
        user_blocks.append({
            "type": "text",
            "text": f"Image/scan from file: {fname}. Extract all PO line items.",
        })
        user_blocks.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/jpeg",
                "data": b64_data,
            },
        })

    if not user_blocks:
        return "[]"

    response = client.messages.create(
        model=DEFAULT_MODEL,
        max_tokens=8192,
        system=[
            {
                "type": "text",
                "text": SYSTEM_PROMPT,
                "cache_control": {"type": "ephemeral"},
            }
        ],
        messages=[{"role": "user", "content": user_blocks}],
    )

    return response.content[0].text


# ---------------------------------------------------------------------------
# Helper: Parse Claude response into DataFrame
# ---------------------------------------------------------------------------
def _repair_truncated_json(raw: str) -> str:
    """
    If Claude's response was cut off mid-JSON (token limit hit), salvage all
    complete objects by trimming to the last '}' and closing the array.
    """
    raw = raw.strip()
    try:
        json.loads(raw)
        return raw  # already valid
    except json.JSONDecodeError:
        pass

    # Find the last complete JSON object and close the array there
    last_brace = raw.rfind('}')
    if last_brace != -1:
        candidate = raw[:last_brace + 1] + ']'
        # Clean up trailing comma before the closing bracket
        candidate = re.sub(r',\s*\]$', ']', candidate)
        try:
            json.loads(candidate)
            return candidate
        except json.JSONDecodeError:
            pass

    return raw  # couldn't repair


def parse_response_to_df(raw: str) -> pd.DataFrame:
    """Parse Claude JSON response into an Odoo-shaped DataFrame."""
    # Strip markdown fences if present
    cleaned = raw.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    # Attempt to repair truncated JSON (happens with very large POs)
    repaired = _repair_truncated_json(cleaned)
    was_truncated = repaired != cleaned and repaired != raw.strip()

    try:
        items = json.loads(repaired)
    except json.JSONDecodeError:
        st.error("Claude devolvió JSON inválido. Respuesta cruda mostrada abajo.")
        st.code(raw)
        return pd.DataFrame(columns=ALL_ODOO_COLUMNS)

    if not isinstance(items, list):
        items = [items]

    if was_truncated:
        st.warning(
            f"⚠️ Este pedido es muy grande — se recuperaron {len(items)} líneas. "
            "Es posible que algunas líneas al final estén incompletas. "
            "Por favor revisa el resultado."
        )

    rows = []
    for item in items:
        rows.append({
            "client_order_ref": str(item.get("client_order_ref", "")),
            "date_order": str(item.get("date_order", "")),
            "commitment_date": str(item.get("commitment_date", "")),
            "*//Cliente": str(item.get("cliente", "")),
            "Producto": str(item.get("producto", "")),
            "Producto_cliente": str(item.get("producto_cliente", "")),
            "Codigo Valido?": "",
            "order_line / product_uom_qty": item.get("qty", 0),
            "id": "",
            "partner_id / id": "",
            "order_line / product_id / id": "",
        })

    df = pd.DataFrame(rows)
    # Ensure all required columns exist (in case Claude omitted some)
    for col in ALL_ODOO_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    # ── Enforce commitment_date = date_order + 2 days when missing/same ──────
    from datetime import timedelta
    for idx, row in df.iterrows():
        d_order = str(row.get("date_order", "")).strip()
        d_commit = str(row.get("commitment_date", "")).strip()
        if d_order and (not d_commit or d_commit == d_order):
            try:
                df.at[idx, "commitment_date"] = (
                    datetime.strptime(d_order, "%Y-%m-%d") + timedelta(days=2)
                ).strftime("%Y-%m-%d")
            except ValueError:
                pass  # date not parseable — leave as-is

    return df


# ---------------------------------------------------------------------------
# Helper: Export to Odoo-ready Excel (uses original template as base)
# ---------------------------------------------------------------------------
def to_odoo_excel(df: pd.DataFrame) -> bytes:
    """
    Write extracted data into the original Odoo template.
    Preserves Clientes and Productos sheets, injects data into
    Import_ventas_LODI and propagates formulas for cols F, H, I, J.
    """
    import openpyxl
    from openpyxl import load_workbook

    template_path = Path(__file__).parent / "Carga Pedidos Odoo.xlsx"

    if template_path.exists():
        wb = load_workbook(str(template_path))
        ws = wb["Import_ventas_LODI"]
        # Clear existing data rows (keep header row 1)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        # Fallback: create fresh workbook with one sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import_ventas_LODI"
        headers = [
            "client_order_ref", "date_order", "commitment_date", "*//Cliente",
            "Producto", "Codigo Valido?", "order_line / product_uom_qty",
            "id", "partner_id / id", "order_line / product_id / id",
        ]
        ws.append(headers)

    # Write extracted rows + formulas
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        # Columns A–D only on the first data row; blank on all subsequent rows
        if i == 2:
            ws[f"A{i}"] = row.get("client_order_ref", "")
            ws[f"B{i}"] = row.get("date_order", "")
            ws[f"C{i}"] = row.get("commitment_date", "")
            ws[f"D{i}"] = row.get("*//Cliente", "")
        producto_val = row.get("Producto", "")
        try:
            producto_val = int(producto_val) if str(producto_val).strip().isdigit() else float(producto_val) if str(producto_val).strip().replace('.','',1).isdigit() else producto_val
        except (ValueError, TypeError):
            pass
        ws[f"E{i}"] = producto_val
        ws[f"G{i}"] = row.get("order_line / product_uom_qty", 0)
        # Formulas for validation columns
        ws[f"F{i}"] = f'=IF(E{i}="","",IF(ISERROR(VLOOKUP(E{i},Productos!B:B,1,FALSE)),"MAL","BIEN"))'
        ws[f"H{i}"] = f'=IF(A{i}="","","import_saleorder_"&A{i}&I{i}&J{i})'
        ws[f"I{i}"] = f'=_xlfn.IFNA(VLOOKUP(D{i},Clientes!A:B,2,0),"")'
        ws[f"J{i}"] = f'=_xlfn.IFNA(VLOOKUP(E{i},Productos!B:D,3,0),"")'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Main UI — Custom CSS
# ---------------------------------------------------------------------------

# Pre-encode logo once (no expand button anywhere)
import base64 as _b64
_logo_path = Path(__file__).parent / "lodi_logo.png"
_logo_b64  = ""
if _logo_path.exists():
    with open(_logo_path, "rb") as _lf:
        _logo_b64 = _b64.b64encode(_lf.read()).decode()
_logo_img_header = (
    f"<img src='data:image/png;base64,{_logo_b64}' "
    "style='width:56px;pointer-events:none;user-select:none;flex-shrink:0'>"
    if _logo_b64 else ""
)
_logo_img_sidebar = (
    f"<img src='data:image/png;base64,{_logo_b64}' "
    "style='width:100px;pointer-events:none;user-select:none;display:block;margin:auto'>"
    if _logo_b64 else ""
)

# ── UI CHANGE: Full CSS rewrite — strict 5-color palette, typography scale, ──
# ── sidebar depth, secondary buttons, pulse CTA, teal download, card system ──
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
html, body, [class*="css"] { font-family: 'Inter', 'Segoe UI', sans-serif !important; }

/* ── Hide Streamlit chrome ───────────────────────────────────────────────── */
#MainMenu, footer, header, [data-testid="stToolbar"] { visibility: hidden; display: none; }

/* ══ COLOR SYSTEM ════════════════════════════════════════════════════════════
   BG0  (page)   : #080C12
   BG1  (sidebar): #0F1620
   BG2  (cards)  : #162030
   BG3  (inputs) : #1C2A3C
   Text primary  : #F0F4F8
   Text secondary: #7A8FA6
   Text muted    : #3D5166
   Border subtle : rgba(255,255,255,0.07)
   Border accent : rgba(232,98,42,0.5)
   Brand primary : #E8622A  (CTAs + key accents only)
   Brand accent  : #f59e0b  (warnings + secondary highlights)
══════════════════════════════════════════════════════════════════════════════ */

/* ══ TYPOGRAPHY SCALE ═════════════════════════════════════════════════════
   Display  : 1.4rem / 800
   Heading  : 1rem   / 700 / uppercase / ls 0.6px
   Body     : 0.875rem / 400
   Small    : 0.78rem  / 500
   Micro    : 0.68rem  / 500 / uppercase / ls 0.5px
══════════════════════════════════════════════════════════════════════════════ */

/* ── TRANSITIONS — all interactive elements ──────────────────────────────── */
* { transition-timing-function: cubic-bezier(0.4, 0, 0.2, 1); }

/* ══ DARK PAGE ═══════════════════════════════════════════════════════════════ */
.stApp {
    background-color: #080C12 !important;
    background-image: radial-gradient(
        ellipse 70% 40% at 65% -5%,
        rgba(232,98,42,0.13) 0%, transparent 55%) !important;
    min-height: 100vh;
}
section[data-testid="stMain"] { background: transparent !important; }
.block-container { padding-top: 1rem !important; padding-bottom: 2.5rem !important; }

/* ══ SIDEBAR ══════════════════════════════════════════════════════════════════
   BG1 (#0F1620) — distinctly deeper than page, warm orange right border      */
section[data-testid="stSidebar"] > div:first-child {
    background: #0F1620 !important;
    border-right: 1px solid rgba(232,98,42,0.15) !important;
    padding: 1.4rem 1rem !important;
}
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div { color: #7A8FA6 !important; }
section[data-testid="stSidebar"] hr {
    border-color: rgba(255,255,255,0.06) !important;
    margin: 0.8rem 0 !important;
}

/* Sidebar section headers — orange-tinted at 60% */
.sb-section-header {
    font-size: 0.68rem !important;
    font-weight: 700 !important;
    color: rgba(232,98,42,0.65) !important;
    text-transform: uppercase !important;
    letter-spacing: 0.8px !important;
    font-family: 'Inter', sans-serif !important;
    margin-bottom: 0.5rem !important;
}

/* Sidebar info card (claude-haiku block) */
.sb-info-card {
    background: #162030;
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 10px;
    padding: 0.75rem 0.9rem;
    margin-top: 0.5rem;
}
.sb-info-row {
    display: flex;
    align-items: center;
    gap: 0.5rem;
    font-size: 0.78rem;
    color: #7A8FA6;
    padding: 0.18rem 0;
    font-family: 'Inter', sans-serif;
}
.sb-info-row .accent { color: #E8622A !important; font-weight: 600 !important; }

/* Sidebar buttons — default destructive red tint */
section[data-testid="stSidebar"] .stButton > button {
    background: rgba(239,68,68,0.1) !important;
    border: 1px solid rgba(239,68,68,0.3) !important;
    color: #fca5a5 !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.875rem !important;
    transition: all 0.18s cubic-bezier(0.4,0,0.2,1) !important;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    background: #dc2626 !important;
    border-color: #dc2626 !important;
    color: white !important;
    transform: none !important;
}

/* ══ LOGOUT BUTTON — subtle text-link style (overrides sidebar default) ════ */
section[data-testid="stSidebar"] .st-key-logout_btn button {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    color: #3D5166 !important;
    font-size: 0.75rem !important;
    font-weight: 500 !important;
    text-decoration: underline !important;
    text-decoration-color: rgba(61,81,102,0.4) !important;
    border-radius: 4px !important;
    padding: 4px 8px !important;
}
section[data-testid="stSidebar"] .st-key-logout_btn button:hover {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    color: #E8622A !important;
    text-decoration-color: rgba(232,98,42,0.4) !important;
    transform: none !important;
}

/* Sidebar API key input */
section[data-testid="stSidebar"] [data-testid="stTextInput"] > div > div > input {
    background-color: #1C2A3C !important;
    border: 1px solid #2A3F57 !important;
    border-radius: 10px !important;
    color: #F0F4F8 !important;
    -webkit-text-fill-color: #F0F4F8 !important;
}

/* ══ NUEVA PO BUTTON — secondary ghost style ══════════════════════════════ */
.st-key-new_po_btn button {
    background: transparent !important;
    border: 1px solid rgba(255,255,255,0.18) !important;
    color: #7A8FA6 !important;
    border-radius: 99px !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    padding: 6px 14px !important;
    box-shadow: none !important;
    white-space: nowrap !important;
    letter-spacing: 0.2px !important;
    transition: all 0.18s cubic-bezier(0.4,0,0.2,1) !important;
}
.st-key-new_po_btn button:hover {
    border-color: #E8622A !important;
    color: #E8622A !important;
    background: rgba(232,98,42,0.06) !important;
    transform: none !important;
    box-shadow: none !important;
}

/* ══ LANGUAGE TOGGLE — small pill, auto-width ════════════════════════════ */
.st-key-lang_main button {
    background: rgba(255,255,255,0.06) !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    color: #7A8FA6 !important;
    border-radius: 99px !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    padding: 6px 14px !important;
    box-shadow: none !important;
    white-space: nowrap !important;
    letter-spacing: 0.2px !important;
    transition: all 0.18s cubic-bezier(0.4,0,0.2,1) !important;
}
.st-key-lang_main button:hover {
    background: #E8622A !important;
    border-color: #E8622A !important;
    color: white !important;
    transform: none !important;
    opacity: 1 !important;
}

/* ══ HEADER CARD ═════════════════════════════════════════════════════════════
   BG2 card, 5px orange left-border, stronger shadow                          */
.lodi-topbar {
    display: flex;
    align-items: center;
    gap: 1.2rem;
    background: #162030;
    border: 1px solid rgba(255,255,255,0.07);
    border-left: 5px solid #E8622A;
    border-radius: 16px;
    padding: 1.1rem 1.6rem;
    margin-bottom: 1.2rem;
    box-shadow: 0 0 0 1px rgba(232,98,42,0.12), 0 8px 32px rgba(0,0,0,0.4);
}
.lodi-topbar h1 {
    font-size: 1.4rem;
    font-weight: 800;
    color: #F0F4F8;
    margin: 0;
    line-height: 1.2;
    font-family: 'Inter', sans-serif;
    letter-spacing: -0.3px;
}
.lodi-topbar p {
    font-size: 0.78rem;
    color: #7A8FA6;
    margin: 0.2rem 0 0;
    font-family: 'Inter', sans-serif;
}

/* ══ STEPPER ══════════════════════════════════════════════════════════════════
   3-step horizontal progress indicator                                        */
.lodi-stepper {
    display: flex;
    align-items: flex-start;
    justify-content: center;
    gap: 0;
    margin: 0.6rem 0 1.2rem;
    font-family: 'Inter', sans-serif;
}
.step-item {
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 0.35rem;
    min-width: 90px;
}
.step-circle {
    width: 30px;
    height: 30px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.75rem;
    font-weight: 700;
    background: #1C2A3C;
    border: 2px solid #2A3F57;
    color: #3D5166;
    flex-shrink: 0;
    transition: all 0.3s ease;
}
.step-circle.active {
    background: #E8622A;
    border-color: #E8622A;
    color: white;
    box-shadow: 0 0 0 4px rgba(232,98,42,0.18);
}
.step-circle.done {
    background: rgba(232,98,42,0.15);
    border-color: rgba(232,98,42,0.5);
    color: #E8622A;
}
.step-label {
    font-size: 0.68rem;
    font-weight: 600;
    color: #3D5166;
    text-align: center;
    letter-spacing: 0.2px;
    white-space: nowrap;
}
.step-label.active { color: #E8622A; }
.step-label.done   { color: #7A8FA6; }
.step-connector {
    height: 2px;
    width: 60px;
    background: #1C2A3C;
    border-radius: 1px;
    margin-top: 14px;
    flex-shrink: 0;
    transition: background 0.3s ease;
}
.step-connector.done { background: rgba(232,98,42,0.45); }

/* ══ UPLOAD CARD ═════════════════════════════════════════════════════════════
   Branded dropzone with hover glow, orange Browse button                     */
[data-testid="stFileUploader"] {
    background: transparent !important;
    border: none !important;
    border-radius: 0 !important;
    padding: 0 !important;
    box-shadow: none !important;
}
[data-testid$="Dropzone"] {
    background: rgba(232,98,42,0.04) !important;
    border: 2px dashed rgba(232,98,42,0.4) !important;
    border-radius: 14px !important;
    min-height: 160px !important;
    display: flex !important;
    flex-direction: column !important;
    justify-content: center !important;
    align-items: center !important;
    gap: 0.4rem !important;
    padding: 2.5rem 1.5rem !important;
    transition: all 0.2s cubic-bezier(0.4,0,0.2,1) !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.3), 0 4px 16px rgba(0,0,0,0.2) !important;
}
[data-testid$="Dropzone"]:hover {
    background: rgba(232,98,42,0.08) !important;
    border-color: rgba(232,98,42,0.7) !important;
    box-shadow: 0 0 0 4px rgba(232,98,42,0.08),
                0 1px 3px rgba(0,0,0,0.3), 0 4px 16px rgba(0,0,0,0.2) !important;
}
[data-testid$="Dropzone"] svg {
    color: #E8622A !important;
    width: 2.2rem !important;
    height: 2.2rem !important;
    opacity: 1 !important;
}
[data-testid$="Dropzone"] span, [data-testid$="Dropzone"] p {
    color: #F0F4F8 !important;
    font-size: 0.875rem !important;
    font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
}
[data-testid$="Dropzone"] small {
    color: #7A8FA6 !important;
    font-size: 0.78rem !important;
    font-weight: 500 !important;
    font-family: 'Inter', sans-serif !important;
}
[data-testid$="Dropzone"] button {
    background: rgba(232,98,42,0.15) !important;
    border: 1px solid rgba(232,98,42,0.4) !important;
    color: #E8622A !important;
    border-radius: 8px !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
    padding: 6px 18px !important;
    margin-top: 0.5rem !important;
    transition: all 0.18s cubic-bezier(0.4,0,0.2,1) !important;
    box-shadow: none !important;
}
[data-testid$="Dropzone"] button:hover {
    background: #E8622A !important;
    border-color: #E8622A !important;
    color: white !important;
    transform: none !important;
}

/* ══ PROCESS BUTTON — orange gradient CTA with pulse animation ═══════════ */
@keyframes pulse-glow {
    0%, 100% { box-shadow: 0 4px 20px rgba(232,98,42,0.35); }
    50%       { box-shadow: 0 6px 32px rgba(232,98,42,0.65); }
}
/* Default button style — orange gradient for all main buttons */
div.stButton > button {
    background: linear-gradient(135deg, #E8622A 0%, #f59e0b 100%) !important;
    border: none !important;
    color: white !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    font-family: 'Inter', sans-serif !important;
    border-radius: 12px !important;
    padding: 0.72rem !important;
    box-shadow: 0 4px 20px rgba(232,98,42,0.35) !important;
    letter-spacing: 0.3px !important;
    transition: opacity 0.18s, transform 0.15s, box-shadow 0.18s !important;
}
div.stButton > button:hover {
    opacity: 0.9 !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 24px rgba(232,98,42,0.5) !important;
}
/* Process button — pulse animation */
.st-key-proc_btn button {
    animation: pulse-glow 2s ease-in-out infinite !important;
}
.st-key-proc_btn button:hover {
    animation: none !important;
    opacity: 0.9 !important;
    transform: translateY(-1px) !important;
}

/* ══ DOWNLOAD BUTTON — teal/cyan (success, distinct from orange CTA) ═══════ */
div.stDownloadButton > button {
    background: linear-gradient(135deg, #0891b2 0%, #0e7490 100%) !important;
    border: none !important;
    color: white !important;
    font-weight: 700 !important;
    font-size: 0.95rem !important;
    font-family: 'Inter', sans-serif !important;
    border-radius: 12px !important;
    padding: 0.7rem !important;
    box-shadow: 0 4px 14px rgba(8,145,178,0.3) !important;
    letter-spacing: 0.3px !important;
    transition: opacity 0.18s !important;
}
div.stDownloadButton > button:hover {
    opacity: 0.88 !important;
    transform: none !important;
    box-shadow: 0 6px 20px rgba(8,145,178,0.45) !important;
}

/* ══ RESULTS HEADER — matches topbar exactly ════════════════════════════════ */
.results-header {
    display: flex;
    align-items: center;
    gap: 0.6rem;
    background: #162030;
    border: 1px solid rgba(255,255,255,0.07);
    border-left: 5px solid #E8622A;
    border-radius: 16px;
    padding: 1rem 1.4rem;
    margin: 1.5rem 0 0.5rem;
    font-size: 1rem;
    font-weight: 700;
    color: #F0F4F8;
    font-family: 'Inter', sans-serif;
    box-shadow: 0 0 0 1px rgba(232,98,42,0.12), 0 8px 32px rgba(0,0,0,0.4);
}
/* Row count pill badge */
.count-badge {
    background: rgba(232,98,42,0.1);
    color: #E8622A;
    border-radius: 99px;
    padding: 2px 10px;
    font-size: 0.75rem;
    font-weight: 500;
    font-family: 'Inter', sans-serif;
    border: 1px solid rgba(232,98,42,0.2);
}
.results-sub {
    font-size: 0.78rem;
    color: #7A8FA6;
    margin-bottom: 0.8rem;
    font-family: 'Inter', sans-serif;
    font-weight: 500;
}

/* ══ FORMAT PILLS — branded orange tint ═══════════════════════════════════ */
.pill {
    display: inline-block;
    background: rgba(232,98,42,0.12);
    color: rgba(232,98,42,0.9);
    border: 1px solid rgba(232,98,42,0.3);
    font-size: 0.68rem;
    font-weight: 600;
    padding: 0.2rem 0.55rem;
    border-radius: 6px;
    margin: 0.15rem 0.1rem;
    font-family: 'SF Mono', 'Fira Code', monospace;
}
.pills-row {
    margin: 0.5rem 0 0.8rem;
    display: flex;
    align-items: center;
    gap: 0.2rem;
    flex-wrap: wrap;
}
.pills-row-label {
    font-size: 0.68rem;
    font-weight: 600;
    color: #3D5166;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    font-family: 'Inter', sans-serif;
    margin-right: 0.3rem;
    white-space: nowrap;
}

/* ══ SECTION LABELS — consistent micro heading style ═══════════════════════ */
.section-label {
    font-size: 0.78rem;
    font-weight: 700;
    color: #7A8FA6;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    font-family: 'Inter', sans-serif;
    margin-bottom: 0.8rem;
}
.upload-label {
    font-size: 0.78rem;
    font-weight: 700;
    color: #7A8FA6;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    font-family: 'Inter', sans-serif;
    margin-bottom: 0.5rem;
}

/* ══ TEXT INPUTS (dark theme — main page only, not login) ═══════════════════ */
section[data-testid="stMain"] [data-testid="stTextInput"] label {
    color: #7A8FA6 !important;
    font-size: 0.78rem !important;
    font-weight: 600 !important;
    font-family: 'Inter', sans-serif !important;
    letter-spacing: 0.3px !important;
}
section[data-testid="stMain"] [data-testid="stTextInput"] > div > div > input {
    background-color: #1C2A3C !important;
    border: 1px solid #2A3F57 !important;
    border-radius: 10px !important;
    color: #F0F4F8 !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 0.875rem !important;
    -webkit-text-fill-color: #F0F4F8 !important;
    transition: border-color 0.18s, box-shadow 0.18s !important;
}
section[data-testid="stMain"] [data-testid="stTextInput"] > div > div > input::placeholder {
    color: #3D5166 !important;
    -webkit-text-fill-color: #3D5166 !important;
    opacity: 1 !important;
}
section[data-testid="stMain"] [data-testid="stTextInput"] > div > div > input:focus {
    border-color: #E8622A !important;
    box-shadow: 0 0 0 3px rgba(232,98,42,0.15) !important;
    background-color: #1e3148 !important;
    -webkit-text-fill-color: #F0F4F8 !important;
}

/* ══ PROGRESS BAR ════════════════════════════════════════════════════════════ */
.stProgress > div > div > div > div { background-color: #E8622A !important; }

/* ══ DATA EDITOR — rounded container clips inner table borders ══════════════ */
[data-testid="stDataEditor"] {
    border-radius: 12px !important;
    overflow: hidden !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.3), 0 4px 16px rgba(0,0,0,0.2) !important;
}

/* ══ EXPANDER — card look, branded ══════════════════════════════════════════ */
[data-testid="stExpander"] {
    background: #162030 !important;
    border: 1px solid rgba(255,255,255,0.07) !important;
    border-radius: 12px !important;
}
[data-testid="stExpander"] summary {
    color: #7A8FA6 !important;
    font-size: 0.875rem !important;
    font-weight: 600 !important;
}

/* ══ SECTION DIVIDERS ════════════════════════════════════════════════════════ */
.section-divider {
    height: 1px;
    background: rgba(255,255,255,0.06);
    margin: 1.2rem 0;
    border: none;
}
</style>
""", unsafe_allow_html=True)

# ── Item 1: Spanish dropzone text override (ES mode only) ────────────────────
if st.session_state.lang == "es":
    st.markdown("""
    <style>
    /* Hide native English dropzone text and inject Spanish */
    [data-testid$="Dropzone"] > span:first-of-type {
        font-size: 0 !important;
        line-height: 0 !important;
        display: block !important;
        height: 1.2em !important;
        position: relative !important;
    }
    [data-testid$="Dropzone"] > span:first-of-type::before {
        content: "Arrastra y suelta archivos aquí";
        font-size: 0.875rem !important;
        font-weight: 600 !important;
        color: #F0F4F8 !important;
        font-family: 'Inter', sans-serif !important;
        position: absolute !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
        white-space: nowrap !important;
        line-height: 1.2em !important;
    }
    [data-testid$="Dropzone"] > small {
        font-size: 0 !important;
        line-height: 0 !important;
        display: block !important;
        height: 1.1em !important;
        position: relative !important;
    }
    [data-testid$="Dropzone"] > small::before {
        content: "Máx. 200 MB · XLSX, XLS, PDF, PNG, JPG";
        font-size: 0.78rem !important;
        font-weight: 500 !important;
        color: #7A8FA6 !important;
        font-family: 'Inter', sans-serif !important;
        position: absolute !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
        white-space: nowrap !important;
        line-height: 1.1em !important;
    }
    [data-testid$="Dropzone"] button {
        color: transparent !important;
        font-size: 0 !important;
        position: relative !important;
        min-width: 130px !important;
    }
    [data-testid$="Dropzone"] button::before {
        content: "Seleccionar archivos";
        color: #E8622A !important;
        font-size: 0.82rem !important;
        font-weight: 600 !important;
        font-family: 'Inter', sans-serif !important;
        position: absolute !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
        white-space: nowrap !important;
    }
    [data-testid$="Dropzone"] button:hover::before { color: white !important; }
    </style>
    """, unsafe_allow_html=True)

# ── UI CHANGE: Sidebar — distinct BG1 background, orange underline on logo, ──
# ── branded section headers, info card rows, subtle logout text-link ──────────
with st.sidebar:
    st.markdown(
        f"<div style='text-align:center;padding:0.5rem 0 0.3rem'>"
        f"{_logo_img_sidebar}"
        f"<div style='width:36px;height:2px;background:#E8622A;margin:0.55rem auto 0.3rem;border-radius:1px'></div>"
        f"<div style='font-size:0.68rem;color:#7A8FA6;letter-spacing:1.2px;font-weight:700;"
        f"text-transform:uppercase;font-family:Inter,sans-serif'>Manufacturing</div>"
        f"</div>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    try:
        _has_key = bool(st.secrets.get("ANTHROPIC_API_KEY", ""))
    except Exception:
        _has_key = False
    if not _has_key:
        st.markdown(
            f"<div class='sb-section-header'>{T['api_label']}</div>",
            unsafe_allow_html=True,
        )
        api_key_input = st.text_input(
            "key", type="password", placeholder="sk-ant-api03-...",
            value=st.session_state.get("manual_api_key", ""),
            label_visibility="collapsed",
        )
        if api_key_input:
            st.session_state["manual_api_key"] = api_key_input
            st.success(T["api_active"])
        st.markdown("---")

    st.markdown(
        f"<div class='sb-section-header'>{T['sb_tool']}</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<div style='font-size:0.78rem;color:#7A8FA6;line-height:1.65;font-family:Inter,sans-serif'>"
        f"{T['sb_tool_desc']}</div>",
        unsafe_allow_html=True,
    )
    # ── UI CHANGE: Info card — proper rows instead of <br> line breaks ────────
    st.markdown(
        "<div class='sb-info-card'>"
        "<div class='sb-info-row'>🤖&nbsp;<span class='accent'>claude-haiku</span></div>"
        "<div class='sb-info-row'>🌐&nbsp;Español · English</div>"
        "<div class='sb-info-row'>⚡&nbsp;Ultra-low token cost</div>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.markdown("---")

    # ── Item 9: Logout as subtle text-link ───────────────────────────────────
    if st.button(T["logout_btn"], key="logout_btn", use_container_width=False):
        st.session_state.authenticated = False
        st.session_state.pop("result_df", None)
        st.rerun()

    # ── UI CHANGE: Footer — text-muted #3D5166, intentionally subtle ─────────
    st.markdown(
        "<div style='text-align:center;font-size:0.68rem;color:#3D5166;"
        "margin-top:1rem;font-family:Inter,sans-serif;letter-spacing:0.4px;"
        "text-transform:uppercase'>LODI Manufacturing<br>Monterrey, MX</div>",
        unsafe_allow_html=True,
    )

# ── UI CHANGE: Header row — Nueva PO is secondary ghost pill (st-key CSS), ───
# ── language toggle is auto-width pill (st-key CSS), no full-width orange ─────
_hcol, _ncol, _lcol = st.columns([6, 1, 1])
with _ncol:
    if st.button(T["new_po_btn"], key="new_po_btn", use_container_width=True):
        for _k in ["result_df", "raw_response", "ov_cliente", "ov_ref", "ov_date", "ov_commit"]:
            st.session_state.pop(_k, None)
        st.session_state["ov_reset"] = st.session_state.get("ov_reset", 0) + 1
        st.rerun()
with _lcol:
    if st.button(T["toggle_btn"], key="lang_main", use_container_width=True):
        st.session_state.lang = "en" if st.session_state.lang == "es" else "es"
        st.rerun()
with _hcol:
    st.markdown(
        f"<div class='lodi-topbar'>"
        f"{_logo_img_header}"
        f"<div style='flex:1'><h1>{T['header_title']}</h1><p>{T['header_sub']}</p></div>"
        f"</div>",
        unsafe_allow_html=True,
    )

# ── UPLOAD ────────────────────────────────────────────────────────────────
_r = st.session_state.get("ov_reset", 0)
st.markdown(f"<div class='upload-label'>{T['upload_label']}</div>", unsafe_allow_html=True)
uploaded_files = st.file_uploader(
    "drop",
    type=["xlsx", "xls", "pdf", "png", "jpg", "jpeg"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    key=f"file_uploader_{_r}",
)

if uploaded_files:
    # ── UI CHANGE: files-selected uses secondary text color ──────────────────
    st.markdown(
        f"<div style='font-size:0.78rem;color:#7A8FA6;margin:0.4rem 0 0.2rem;"
        f"font-family:Inter,sans-serif;font-weight:500'>"
        f"✅ {T['files_selected'].format(n=len(uploaded_files))}</div>",
        unsafe_allow_html=True,
    )

# ── Item 2: Format pills below upload zone ────────────────────────────────────
st.markdown(
    f"<div class='pills-row'>"
    f"<span class='pills-row-label'>{T['formats_label']}:</span>"
    f"<span class='pill'>.xlsx</span><span class='pill'>.xls</span>"
    f"<span class='pill'>.pdf</span><span class='pill'>.png</span>"
    f"<span class='pill'>.jpg</span><span class='pill'>.jpeg</span>"
    f"</div>",
    unsafe_allow_html=True,
)

# ── Item 8: 3-step progress stepper ──────────────────────────────────────────
if "result_df" in st.session_state:
    _step = 3
elif uploaded_files:
    _step = 2
else:
    _step = 1

def _render_stepper(step: int, labels: list[str]) -> str:
    parts = []
    for n, lbl in enumerate(labels, 1):
        circ_cls = "done" if n < step else ("active" if n == step else "")
        lbl_cls  = circ_cls
        inner    = "✓" if n < step else str(n)
        parts.append(
            f"<div class='step-item'>"
            f"<div class='step-circle {circ_cls}'>{inner}</div>"
            f"<div class='step-label {lbl_cls}'>{lbl}</div>"
            f"</div>"
        )
        if n < len(labels):
            conn_cls = "done" if n < step else ""
            parts.append(f"<div class='step-connector {conn_cls}'></div>")
    return "<div class='lodi-stepper'>" + "".join(parts) + "</div>"

st.markdown(
    _render_stepper(_step, [T["step1"], T["step2"], T["step3"]]),
    unsafe_allow_html=True,
)

# ── Item 3: Override section (renamed title) ──────────────────────────────────
st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
st.markdown(
    f"<div style='background:#162030;border:1px solid rgba(255,255,255,0.07);"
    f"border-radius:14px;padding:1.4rem;margin:0.4rem 0;box-shadow:0 1px 3px rgba(0,0,0,0.3),0 4px 16px rgba(0,0,0,0.2)'>"
    f"<div class='section-label'>{T['override_title']}</div>"
    f"<div style='background:rgba(245,158,11,0.08);border-left:3px solid #f59e0b;"
    f"border-radius:0 8px 8px 0;padding:0.5rem 0.8rem;margin-bottom:0.9rem;"
    f"font-size:0.78rem;color:#fde68a;font-family:Inter,sans-serif;font-weight:500'>"
    + (
        "⚠️ <strong>Solo llena estos campos si el PO no los incluye.</strong> "
        "Si el PO ya los tiene, déjalos en blanco — la IA los detectará automáticamente."
        if st.session_state.lang == "es" else
        "⚠️ <strong>Only fill these if the PO does not include them.</strong> "
        "If the PO already has them, leave blank — the AI will detect them automatically."
    ) +
    f"</div></div>",
    unsafe_allow_html=True,
)
ov_col1, ov_col2 = st.columns(2)
with ov_col1:
    override_cliente = st.text_input(T["override_cliente"], key=f"ov_cliente_{_r}", placeholder="Ej: RUSAL SA DE CV")
    override_date    = st.text_input(T["override_date"],    key=f"ov_date_{_r}",    placeholder="2026-03-04")
with ov_col2:
    override_ref     = st.text_input(T["override_ref"],     key=f"ov_ref_{_r}",     placeholder="Ej: PO-12345")
    override_commit  = st.text_input(T["override_commit"],  key=f"ov_commit_{_r}",  placeholder="2026-03-15")

# ── UI CHANGE: Process button — centered column wrapper + key for pulse CSS ──
st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
_proc_l, _proc_m, _proc_r = st.columns([1, 3, 1])
with _proc_m:
    _proc_clicked = bool(uploaded_files) and st.button(
        T["process_btn"], key="proc_btn", use_container_width=True,
    )

# ── Item 4: Multi-step loading with st.status() ───────────────────────────────
if _proc_clicked:
    text_parts: list[str] = []
    vision_images: list[tuple[str, str]] = []
    result_df = pd.DataFrame()
    raw_response = ""

    with st.status(T["prog_step1"], expanded=True) as _status:
        _inner_prog = st.progress(0)

        for i, f in enumerate(uploaded_files):
            file_bytes = f.read()
            ext = Path(f.name).suffix.lower()
            _status.update(label=T["prog_file"].format(f=f.name))
            _inner_prog.progress((i + 1) / (len(uploaded_files) + 1))

            if ext in (".xlsx", ".xls"):
                text_parts.append(extract_text_from_excel(file_bytes, f.name))
            elif ext == ".pdf":
                text, needs_vision = extract_text_from_pdf(file_bytes, f.name)
                if needs_vision:
                    # Scanned PDF — vision only
                    try:
                        vision_images.extend(render_pdf_for_vision(file_bytes, f.name))
                    except Exception:
                        st.warning(T["warn_pdf"].format(f=f.name))
                else:
                    # Digital PDF — send text layer AND rendered images
                    text_parts.append(text)
                    try:
                        vision_images.extend(render_pdf_for_vision(file_bytes, f.name))
                    except Exception:
                        pass  # Vision is supplemental; text extraction succeeded
            elif ext in (".png", ".jpg", ".jpeg"):
                text, needs_vision = extract_text_from_image(file_bytes, f.name)
                if needs_vision:
                    vision_images.append((resize_image_for_vision(file_bytes), f.name))
                else:
                    text_parts.append(text)

        _status.update(label=T["prog_step2"])
        _inner_prog.progress(0.85)
        raw_response = call_claude("\n\n".join(text_parts), vision_images)

        # ── Parse response ─────────────────────────────────────────────────
        result_df = parse_response_to_df(raw_response)

        # ── Apply manual overrides ─────────────────────────────────────────
        _r_cur = st.session_state.get("ov_reset", 0)
        ov_cliente = st.session_state.get(f"ov_cliente_{_r_cur}", "").strip()
        ov_ref     = st.session_state.get(f"ov_ref_{_r_cur}",     "").strip()
        ov_date    = st.session_state.get(f"ov_date_{_r_cur}",    "").strip()
        ov_commit  = st.session_state.get(f"ov_commit_{_r_cur}",  "").strip()
        if not result_df.empty:
            if ov_cliente:
                result_df["*//Cliente"] = ov_cliente
            if ov_ref:
                result_df["client_order_ref"] = ov_ref
            if ov_date:
                result_df["date_order"] = ov_date
            if ov_commit:
                result_df["commitment_date"] = ov_commit

        # ── Client name resolution ─────────────────────────────────────────
        api_key = (
            st.session_state.get("manual_api_key", "").strip()
            or os.environ.get("ANTHROPIC_API_KEY", "")
            or st.secrets.get("ANTHROPIC_API_KEY", "")
        )
        if api_key and not result_df.empty:
            unique_clients = result_df["*//Cliente"].dropna().unique()
            client_map: dict[str, str] = {}
            for raw in unique_clients:
                if raw.strip():
                    client_map[raw] = resolve_client_name(raw, api_key)
            result_df["*//Cliente"] = result_df["*//Cliente"].map(
                lambda x: client_map.get(x, x)
            )

        # Ensure all Odoo columns exist; keep Producto_cliente as internal column
        for col in ALL_ODOO_COLUMNS:
            if col not in result_df.columns:
                result_df[col] = ""
        if "Producto_cliente" not in result_df.columns:
            result_df["Producto_cliente"] = ""

        _inner_prog.progress(1.0)
        _status.update(label=T["prog_step3"], state="complete", expanded=False)

    st.session_state["result_df"] = result_df
    st.session_state["raw_response"] = raw_response

# ── RESULTS ───────────────────────────────────────────────────────────────
if "result_df" in st.session_state:
    _stored_df = st.session_state["result_df"]
    n_rows = len(_stored_df)

    # ── UI CHANGE: Results header matches topbar style; count uses pill badge ─
    st.markdown(
        f"<div class='results-header'>{T['results_title']}"
        f"<span class='count-badge'>{n_rows} {T['lines_suffix']}</span></div>",
        unsafe_allow_html=True,
    )
    st.markdown(f"<div class='results-sub'>{T['results_sub']}</div>", unsafe_allow_html=True)

    # ── Dual-code detector ──────────────────────────────────────────────────
    _has_dual = (
        "Producto_cliente" in _stored_df.columns
        and _stored_df["Producto_cliente"].str.strip().ne("").any()
    )

    # ── Item 7: Show sample values in dual-code selector labels ──────────────
    if _has_dual:
        _sample_lodi = ""
        _sample_cli  = ""
        for _, _row in _stored_df.iterrows():
            if not _sample_lodi and str(_row.get("Producto", "")).strip():
                _sample_lodi = str(_row["Producto"]).strip()
            if not _sample_cli and str(_row.get("Producto_cliente", "")).strip():
                _sample_cli = str(_row["Producto_cliente"]).strip()
            if _sample_lodi and _sample_cli:
                break

        _col_a_label = f"A: {_sample_lodi}" if _sample_lodi else T["code_lodi"]
        _col_b_label = f"B: {_sample_cli}"  if _sample_cli  else T["code_cliente"]

        _code_choice = st.radio(
            T["code_choice_label"],
            options=["lodi", "cliente"],
            format_func=lambda x: _col_a_label if x == "lodi" else _col_b_label,
            horizontal=True,
            key="code_choice",
        )
    else:
        _code_choice = "lodi"

    # Build display df: swap Producto when client code is selected
    _display_df = _stored_df.copy()
    if _code_choice == "cliente" and _has_dual:
        _mask = _display_df["Producto_cliente"].str.strip().ne("")
        _display_df.loc[_mask, "Producto"] = _display_df.loc[_mask, "Producto_cliente"]

    # ── Item 5: Use DISPLAY_COLUMNS (hides internal id columns) ──────────────
    for _col in DISPLAY_COLUMNS:
        if _col not in _display_df.columns:
            _display_df[_col] = ""
    _display_df = _display_df[DISPLAY_COLUMNS].copy()

    # ── Item 6: Visual encoding for "Codigo Valido?" column ──────────────────
    def _encode_valid(v: str) -> str:
        v = str(v).strip().upper()
        if v == "BIEN":  return "✅"
        if v == "MAL":   return "❌"
        return "⚪"

    _display_df["Codigo Valido?"] = _display_df["Codigo Valido?"].apply(_encode_valid)

    edited_df = st.data_editor(
        _display_df,
        use_container_width=True,
        num_rows="dynamic",
        key="po_editor",
        column_config={
            "order_line / product_uom_qty": st.column_config.NumberColumn(T["col_qty"], min_value=0, format="%g"),
            "date_order":        st.column_config.TextColumn(T["col_date"]),
            "commitment_date":   st.column_config.TextColumn(T["col_commit"]),
            "client_order_ref":  st.column_config.TextColumn(T["col_ref"]),
            "*//Cliente":        st.column_config.TextColumn(T["col_client"]),
            "Producto":          st.column_config.TextColumn(T["col_product"]),
            "Codigo Valido?":    st.column_config.TextColumn(
                T["col_valid"],
                disabled=True,
                help="Calculado automáticamente en Excel al descargar" if st.session_state.lang == "es"
                     else "Calculated automatically in Excel upon download",
            ),
        },
    )

    st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)

    # ── Item 10: Expander ABOVE download, both full-width ─────────────────────
    with st.expander(T["raw_expander"]):
        st.code(st.session_state.get("raw_response", ""), language="json")

    st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

    # Download button — full width, below expander
    # Reconstruct full Odoo column set from edited display df + stored internal cols
    _export_df = edited_df.copy()
    # Add internal id columns back (always empty — filled by Excel formulas)
    for col in ALL_ODOO_COLUMNS:
        if col not in _export_df.columns:
            _export_df[col] = ""
    _export_df = _export_df[ALL_ODOO_COLUMNS]
    excel_bytes = to_odoo_excel(_export_df)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    st.download_button(
        label=T["download_btn"],
        data=excel_bytes,
        file_name=f"Carga_Pedidos_Odoo_{timestamp}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
